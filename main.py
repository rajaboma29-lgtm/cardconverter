from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
import pdfplumber
import csv, io, re, uuid, os, json, threading
from typing import Dict, Any, List
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from openpyxl import load_workbook

DONATION_LINE_AR = "هذا العمل صدقة جارية عن روح والدتي وأخي رحمهم الله"
ENG_LINE = "Eng : RajabSukker"

app = FastAPI(title="Card Factory")

TMP = os.environ.get("CARD_FACTORY_TMP", "/tmp/card_factory")
os.makedirs(TMP, exist_ok=True)

USER_RE = re.compile(r"Username\s+(\S+)", re.IGNORECASE)
PASS_RE = re.compile(r"Password\s+(\S+)", re.IGNORECASE)
PASS_NEXTLINE_RE = re.compile(r"Password\s*\n\s*([0-9A-Za-z]+)", re.IGNORECASE)

JOBS: Dict[str, Dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()

def _hex_to_rgb(hex_color: str):
    try:
        s = (hex_color or "").strip()
        if s.startswith("#"): s = s[1:]
        if len(s) == 3: s = "".join([c*2 for c in s])
        return (int(s[0:2],16)/255.0, int(s[2:4],16)/255.0, int(s[4:6],16)/255.0)
    except Exception:
        return (1,1,1)


def _clamp_int(v, mn, mx, default):
    try:
        v = int(v)
    except Exception:
        return default
    return max(mn, min(mx, v))

def extract_pairs_from_pdf(pdf_path: str):
    pairs = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            users = USER_RE.findall(text)
            passes = PASS_NEXTLINE_RE.findall(text) or PASS_RE.findall(text)
            for u, p in zip(users, passes):
                u = (u or "").strip(); p = (p or "").strip()
                if u and p: pairs.append((u,p))
    return pairs

def extract_pairs_from_csv_bytes(csv_bytes: bytes):
    s = csv_bytes.decode("utf-8", errors="ignore")
    f = io.StringIO(s)
    reader = csv.reader(f)
    out = []
    for r in reader:
        if not r: 
            continue
        if len(r) >= 2:
            a = (r[0] or "").strip()
            b = (r[1] or "").strip()
            # تجاهل أي صف هيدر
            if a.lower() in ("username","user","اسم المستخدم") and b.lower() in ("password","pass","كلمة المرور"):
                continue
            if a and b:
                out.append((a, b))
    return out

def extract_pairs_from_xlsx_bytes(xlsx_bytes: bytes):
    bio = io.BytesIO(xlsx_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a = (str(row[0]).strip() if len(row)>0 and row[0] is not None else "")
        b = (str(row[1]).strip() if len(row)>1 and row[1] is not None else "")
        # تجاهل أي صف هيدر (حتى لو مكرر بالمنتصف)
        if a.lower() in ("username","user","اسم المستخدم") and b.lower() in ("password","pass","كلمة المرور"):
            continue
        if a and b:
            out.append((a, b))
    return out

def dedupe(pairs):
    seen = set(); out=[]
    for u,p in pairs:
        if not u or not p: continue
        if u in seen: continue
        seen.add(u); out.append((u,p))
    return out

def draw_text_fit(c, text, x, y, w, h, font="Helvetica", max_size=14, min_size=6, color_hex="#ffffff"):
    text = str(text)
    r,g,b = _hex_to_rgb(color_hex)
    c.setFillColorRGB(r,g,b)
    size = float(max_size)
    c.setFont(font, size)
    while size >= float(min_size) and c.stringWidth(text, font, size) > max(1, w):
        size -= 1
        c.setFont(font, size)
    tx = x + (w - c.stringWidth(text, font, size)) / 2
    ty = y + (h - size) / 2
    c.drawString(tx, ty, text)


def auto_layout(pagesize, cols:int, rows:int):
    """
    اختيار هامش/فراغ تلقائي لتعظيم مساحة البطاقة مع قيود بسيطة.
    """
    W, H = pagesize
    best = None
    # قيود منطقية
    for margin in range(6, 26):        # نقاط
        for gap in range(2, 13):       # نقاط
            cardW = (W - 2*margin - (cols-1)*gap) / max(1, cols)
            cardH = (H - 2*margin - (rows-1)*gap) / max(1, rows)
            if cardW <= 0 or cardH <= 0:
                continue
            # لا نسمح بقيم صغيرة جدًا (تجنب تخريب الشبكات الضخمة)
            if cardW < 20 or cardH < 20:
                continue
            area = cardW * cardH
            # عقوبة بسيطة لهوامش ضخمة (نفضل هوامش أصغر إذا نفس المساحة تقريبًا)
            score = area - (margin*2 + gap*5)
            if best is None or score > best[0]:
                best = (score, margin, gap)
    if best:
        return best[1], best[2]
    return 18, 6

def generate_a4_pdf(bg_image_path, pairs, placement, out_path, cols=5, rows=10, margin_pt=18, gap_pt=6, progress_cb=None, progress_base=40, progress_span=60, orientation='portrait', crop_marks: bool = False, auto_fit: bool = False):
    pagesize = landscape(A4) if (orientation or '').lower().startswith('land') else A4
    W,H = pagesize

    if auto_fit:
        margin_pt, gap_pt = auto_layout(pagesize, cols, rows)
    cardW = (W - 2*margin_pt - (cols-1)*gap_pt) / cols
    cardH = (H - 2*margin_pt - (rows-1)*gap_pt) / rows

    bg = ImageReader(bg_image_path)
    c = canvas.Canvas(out_path, pagesize=pagesize)

    total = len(pairs) or 1
    footer = f"{DONATION_LINE_AR} — {ENG_LINE}"

    ubox = placement["username"]; pbox = placement["password"]
    u_size = float(ubox.get("font_size", 14)); p_size = float(pbox.get("font_size", 14))
    u_color = ubox.get("color", "#ffffff"); p_color = pbox.get("color", "#ffffff")

    def draw_crop_marks(x0, y0, cardW, cardH, L=6):
        # خطوط صغيرة عند زوايا البطاقة (داخل وخارج قليلًا)
        c.setStrokeColorRGB(0.85,0.85,0.85)
        c.setLineWidth(0.3)
        # top-left
        c.line(x0, y0+cardH, x0+L, y0+cardH)
        c.line(x0, y0+cardH, x0, y0+cardH-L)
        # top-right
        c.line(x0+cardW, y0+cardH, x0+cardW-L, y0+cardH)
        c.line(x0+cardW, y0+cardH, x0+cardW, y0+cardH-L)
        # bottom-left
        c.line(x0, y0, x0+L, y0)
        c.line(x0, y0, x0, y0+L)
        # bottom-right
        c.line(x0+cardW, y0, x0+cardW-L, y0)
        c.line(x0+cardW, y0, x0+cardW, y0+L)

    i=0
    while i < len(pairs):
        c.setFillColor(colors.white)
        c.setFont("Helvetica", 7)
        c.drawString(margin_pt, 6, footer)

        for r in range(rows):
            for col in range(cols):
                if i >= len(pairs): break
                u,p = pairs[i]
                x0 = margin_pt + col * (cardW + gap_pt)
                y0 = H - margin_pt - (r+1) * cardH - r * gap_pt
                c.drawImage(bg, x0, y0, width=cardW, height=cardH, mask='auto')
                if crop_marks:
                    draw_crop_marks(x0, y0, cardW, cardH)

                def map_box(box):
                    bx = x0 + box["x"] * cardW
                    by = y0 + (1 - box["y"] - box["h"]) * cardH
                    bw = box["w"] * cardW
                    bh = box["h"] * cardH
                    return bx,by,bw,bh

                ux,uy,uw,uh = map_box(ubox)
                px,py,pw,ph = map_box(pbox)

                draw_text_fit(c, u, ux, uy, uw, uh, max_size=u_size, min_size=6, color_hex=u_color)
                draw_text_fit(c, p, px, py, pw, ph, max_size=p_size, min_size=6, color_hex=p_color)

                i += 1
                if progress_cb:
                    pct = progress_base + (i/total)*progress_span
                    progress_cb(pct)
            if i >= len(pairs): break
        c.showPage()
    c.save()

def _set_job(job_id: str, **kwargs):
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(kwargs)

def _process_job(job_id: str):
    try:
        with JOBS_LOCK:
            j = JOBS[job_id]

        def prog(pct):
            _set_job(job_id, progress=int(max(0, min(100, pct))))

        prog(1)

        # 1) Extract from multiple files
        all_pairs: List[tuple] = []
        files = j["data_files"]  # list of dicts: {name,path}
        total_files = len(files) or 1

        for idx, fmeta in enumerate(files, start=1):
            name = fmeta["name"]
            path = fmeta["path"]
            # 0..40% موزعة على الملفات
            base = ((idx-1)/total_files)*40
            span = (1/total_files)*40

            if name.endswith(".pdf"):
                pairs = extract_pairs_from_pdf(path)
            elif name.endswith(".xlsx"):
                with open(path, "rb") as f:
                    pairs = extract_pairs_from_xlsx_bytes(f.read())
            else:
                with open(path, "rb") as f:
                    pairs = extract_pairs_from_csv_bytes(f.read())

            all_pairs.extend(pairs)
            prog(base + span)

        pairs = dedupe(all_pairs)[:j["max_cards"]]
        count = len(pairs)

        # 2) CSV
        with open(j["out_csv"], "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["username","password"]); w.writerows(pairs)

        # 3) Preview PDF (first page only)
        preview_pairs = pairs[: j["cols"] * j["rows"]]
        generate_a4_pdf(
            j["design_path"], preview_pairs, j["placement"], j["out_preview"],
            cols=j["cols"], rows=j["rows"], margin_pt=j["margin_pt"], gap_pt=j["gap_pt"],
            orientation=j.get("orientation","portrait"),
            crop_marks=j.get("crop_marks", False),
            auto_fit=j.get("auto_fit", False),
            progress_cb=prog, progress_base=40, progress_span=20
        )

                # إذا معاينة فقط: لا تولّد الملف النهائي
        if j.get("preview_only", False):
            _set_job(job_id, status="done", progress=100, count=count)
            return

        # 4) Full PDF
        generate_a4_pdf(
            j["design_path"], pairs, j["placement"], j["out_pdf"],
            cols=j["cols"], rows=j["rows"], margin_pt=j["margin_pt"], gap_pt=j["gap_pt"],
            orientation=j.get("orientation","portrait"),
            crop_marks=j.get("crop_marks", False),
            auto_fit=j.get("auto_fit", False),
            progress_cb=prog, progress_base=60, progress_span=40
        )

        _set_job(job_id, status="done", progress=100, count=count)
    except Exception as e:
        _set_job(job_id, status="error", error=str(e))

@app.get("/", response_class=HTMLResponse)
def index():
    here = os.path.dirname(__file__)
    with open(os.path.join(here, "static", "index.html"), "r", encoding="utf-8") as f:
        return f.read()

@app.post("/api/process")
async def start_process(
    design: UploadFile = File(...),
    datafiles: List[UploadFile] = File(...),
    placement_json: str = Form(...),
    max_cards: int = Form(10000),
    cols: int = Form(5),
    rows: int = Form(10),
    margin_pt: int = Form(18),
    gap_pt: int = Form(6),
    orientation: str = Form('portrait'),
    crop_marks: bool = Form(False),
    auto_fit: bool = Form(False),
    preview_only: bool = Form(False),
):
    job_id = str(uuid.uuid4())

    # حماية من قيم غير منطقية
    cols = _clamp_int(cols, 1, 10, 5)
    rows = _clamp_int(rows, 1, 14, 10)
    max_cards = _clamp_int(max_cards, 1, 10000, 10000)

    ext = os.path.splitext(design.filename or "")[1].lower() or ".png"
    if ext not in (".png",".jpg",".jpeg"):
        return JSONResponse({"error":"يرجى رفع تصميم بصيغة PNG أو JPG."}, status_code=400)

    # validate multiple files types
    allowed = (".pdf",".csv",".xlsx")
    if not datafiles or len(datafiles) == 0:
        return JSONResponse({"error":"ارفع ملف بيانات واحد على الأقل."}, status_code=400)

    design_path = os.path.join(TMP, f"{job_id}_design{ext}")
    with open(design_path, "wb") as f:
        f.write(await design.read())

    saved_files = []
    for df in datafiles:
        name = (df.filename or "").lower()
        if not name.endswith(allowed):
            return JSONResponse({"error":"الملفات المدعومة: PDF / CSV / XLSX فقط."}, status_code=400)
        ext2 = os.path.splitext(name)[1]
        data_path = os.path.join(TMP, f"{job_id}_{len(saved_files)+1}{ext2}")
        b = await df.read()
        with open(data_path, "wb") as f:
            f.write(b)
        saved_files.append({"name": name, "path": data_path})

    try:
        placement = json.loads(placement_json)
        assert "username" in placement and "password" in placement
    except Exception:
        return JSONResponse({"error":"بيانات تحديد الأماكن غير صالحة."}, status_code=400)

    out_pdf = os.path.join(TMP, f"{job_id}_print.pdf")
    out_preview = os.path.join(TMP, f"{job_id}_preview.pdf")
    out_csv = os.path.join(TMP, f"{job_id}_clean.csv")

    with JOBS_LOCK:
        JOBS[job_id] = {
            "status":"processing", "progress":0, "count":0, "error":"",
            "design_path":design_path,
            "data_files": saved_files,
            "placement":placement,
            "max_cards":max_cards,
            "cols":cols, "rows":rows, "margin_pt":margin_pt, "gap_pt":gap_pt,
            "out_pdf":out_pdf, "out_preview": out_preview, "out_csv":out_csv,
            "orientation": (orientation or "portrait"),
            "crop_marks": bool(crop_marks),
            "auto_fit": bool(auto_fit),
            "preview_only": bool(preview_only)
        }

    threading.Thread(target=_process_job, args=(job_id,), daemon=True).start()
    return {"job_id": job_id}

@app.get("/api/status/{job_id}")
def status(job_id: str):
    with JOBS_LOCK:
        j = JOBS.get(job_id)
        if not j:
            return JSONResponse({"error":"المهمة غير موجودة."}, status_code=404)
        resp = {"status": j["status"], "progress": j["progress"], "count": j.get("count",0)}
        if j["status"] == "done":
            count = j.get("count",0)
            resp.update({
                "preview_url": f"/api/download/preview/{job_id}",
                "csv_url": f"/api/download/csv/{job_id}",
                "file_name": f"Abo Omar_{count}",
                "preview_only": bool(j.get("preview_only", False))
            })
            if not j.get("preview_only", False):
                resp["pdf_url"] = f"/api/download/pdf/{job_id}"
        if j["status"] == "error":
            resp["error"] = j.get("error","")
        return resp

@app.get("/api/download/preview/{job_id}")
def dl_preview(job_id: str):
    with JOBS_LOCK:
        j = JOBS.get(job_id)
        if not j or j.get("status") != "done":
            return JSONResponse({"error":"المعاينة غير جاهزة بعد."}, status_code=400)
        count = j.get("count",0)
    return FileResponse(os.path.join(TMP, f"{job_id}_preview.pdf"),
                        media_type="application/pdf",
                        filename=f"Abo Omar_{count}_preview.pdf")

@app.get("/api/download/pdf/{job_id}")
def dl_pdf(job_id: str):
    with JOBS_LOCK:
        j = JOBS.get(job_id)
        if not j or j.get("status") != "done":
            return JSONResponse({"error":"الملف غير جاهز بعد."}, status_code=400)
        count = j.get("count",0)
    return FileResponse(os.path.join(TMP, f"{job_id}_print.pdf"),
                        media_type="application/pdf",
                        filename=f"Abo Omar_{count}.pdf")

@app.get("/api/download/csv/{job_id}")
def dl_csv(job_id: str):
    with JOBS_LOCK:
        j = JOBS.get(job_id)
        if not j or j.get("status") != "done":
            return JSONResponse({"error":"الملف غير جاهز بعد."}, status_code=400)
        count = j.get("count",0)
    return FileResponse(os.path.join(TMP, f"{job_id}_clean.csv"),
                        media_type="text/csv",
                        filename=f"Abo Omar_{count}.csv")
